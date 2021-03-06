# Redo the study using 3 years only
import os
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
import pickle
from concurrent.futures import ThreadPoolExecutor as T
from openpyxl import Workbook, load_workbook

import xlrd
from multiprocessing import Lock


class ProcessData:
    def __init__(self, files):
        self.wd = os.getcwd()
        self.data_dir = self.wd + '\\data\\Potential'
        self.files = files

    def main(self):
        print('Process Started')
        files = self.files.copy()
        print('Start Process 1')
        # Stage 1
        with T(max_workers=6) as pool:
            pool.map(self.process_files, self.files)
        print('Process Done 1')

    def process_files(self, f):
        sr = self.data_dir + '\\' + f + '.xlsx'
        print(sr)
        w = load_workbook(filename=sr, read_only=False)
        df = pd.read_excel(sr)

        # Get the min distance from min and max
        df['d1_mm'] = abs(df['x0_min'] - df['x1_min'])
        df['d2_mm'] = abs(df['x0_min'] - df['x2_min'])
        df['d3_mm'] = abs(df['x0_min'] - df['x3_min'])
        df['d4_mm'] = abs(df['x0_min'] - df['x4_min'])
        df['d5_mm'] = abs(df['x0_min'] - df['x5_min'])
        df['d6_mm'] = abs(df['x0_min'] - df['x6_min'])

        df['d1_xx'] = abs(df['x0_max'] - df['x1_max'])
        df['d2_xx'] = abs(df['x0_max'] - df['x2_max'])
        df['d3_xx'] = abs(df['x0_max'] - df['x3_max'])
        df['d4_xx'] = abs(df['x0_max'] - df['x4_max'])
        df['d5_xx'] = abs(df['x0_max'] - df['x5_max'])
        df['d6_xx'] = abs(df['x0_max'] - df['x6_max'])

        df['d1_mx'] = abs(df['x0_min'] - df['x1_max'])
        df['d2_mx'] = abs(df['x0_min'] - df['x2_max'])
        df['d3_mx'] = abs(df['x0_min'] - df['x3_max'])
        df['d4_mx'] = abs(df['x0_min'] - df['x4_max'])
        df['d5_mx'] = abs(df['x0_min'] - df['x5_max'])
        df['d6_mx'] = abs(df['x0_min'] - df['x6_max'])

        df['d1_xm'] = abs(df['x0_max'] - df['x1_min'])
        df['d2_xm'] = abs(df['x0_max'] - df['x2_min'])
        df['d3_xm'] = abs(df['x0_max'] - df['x3_min'])
        df['d4_xm'] = abs(df['x0_max'] - df['x4_min'])
        df['d5_xm'] = abs(df['x0_max'] - df['x5_min'])
        df['d6_xm'] = abs(df['x0_max'] - df['x6_min'])
        # min of every group
        df['SComb_mm'] = df.iloc[:, 15:21].min(axis=1)
        df['SComb_xx'] = df.iloc[:, 21:27].min(axis=1)
        df['SComb_mx'] = df.iloc[:, 27:33].min(axis=1)
        df['SComb_xm'] = df.iloc[:, 33:39].min(axis=1)
        # ---------------------------------------------
        # how close every variable to the min of all variables?
        df['s1_mm'] = df['d1_mm'] - df['SComb_mm']
        df['s2_mm'] = df['d2_mm'] - df['SComb_mm']
        df['s3_mm'] = df['d3_mm'] - df['SComb_mm']
        df['s4_mm'] = df['d4_mm'] - df['SComb_mm']
        df['s5_mm'] = df['d5_mm'] - df['SComb_mm']
        df['s6_mm'] = df['d6_mm'] - df['SComb_mm']

        df['s1_xx'] = df['d1_xx'] - df['SComb_xx']
        df['s2_xx'] = df['d2_xx'] - df['SComb_xx']
        df['s3_xx'] = df['d3_xx'] - df['SComb_xx']
        df['s4_xx'] = df['d4_xx'] - df['SComb_xx']
        df['s5_xx'] = df['d5_xx'] - df['SComb_xx']
        df['s6_xx'] = df['d6_xx'] - df['SComb_xx']

        df['s1_mx'] = df['d1_mx'] - df['SComb_mx']
        df['s2_mx'] = df['d2_mx'] - df['SComb_mx']
        df['s3_mx'] = df['d3_mx'] - df['SComb_mx']
        df['s4_mx'] = df['d4_mx'] - df['SComb_mx']
        df['s5_mx'] = df['d5_mx'] - df['SComb_mx']
        df['s6_mx'] = df['d6_mx'] - df['SComb_mx']

        df['s1_xm'] = df['d1_xm'] - df['SComb_xm']
        df['s2_xm'] = df['d2_xm'] - df['SComb_xm']
        df['s3_xm'] = df['d3_xm'] - df['SComb_xm']
        df['s4_xm'] = df['d4_xm'] - df['SComb_xm']
        df['s5_xm'] = df['d5_xm'] - df['SComb_xm']
        df['s6_xm'] = df['d6_xm'] - df['SComb_xm']

        wspd = w.create_sheet("Processed_Data", 1)
        for r in dataframe_to_rows(df, index=False, header=True):
            wspd.append(r)

        sim_mm = [1 - df['d1_mm'].mean(),
                  1 - df['d2_mm'].mean(),
                  1 - df['d3_mm'].mean(),
                  1 - df['d4_mm'].mean(),
                  1 - df['d5_mm'].mean(),
                  1 - df['d6_mm'].mean(),
                  1 - df['SComb_mm'].mean()]

        sim_xx = [1 - df['d1_xx'].mean(),
                  1 - df['d2_xx'].mean(),
                  1 - df['d3_xx'].mean(),
                  1 - df['d4_xx'].mean(),
                  1 - df['d5_xx'].mean(),
                  1 - df['d6_xx'].mean(),
                  1 - df['SComb_xx'].mean()]

        sim_mx = [1 - df['d1_mx'].mean(),
                  1 - df['d2_mx'].mean(),
                  1 - df['d3_mx'].mean(),
                  1 - df['d4_mx'].mean(),
                  1 - df['d5_mx'].mean(),
                  1 - df['d6_mx'].mean(),
                  1 - df['SComb_mx'].mean()]

        sim_xm = [1 - df['d1_xm'].mean(),
                  1 - df['d2_xm'].mean(),
                  1 - df['d3_xm'].mean(),
                  1 - df['d4_xm'].mean(),
                  1 - df['d5_xm'].mean(),
                  1 - df['d6_xm'].mean(),
                  1 - df['SComb_xm'].mean()]

        sim_mm_r = [round(x, 5) for x in sim_mm]
        sim_xx_r = [round(x, 5) for x in sim_xx]
        sim_mx_r = [round(x, 5) for x in sim_mx]
        sim_xm_r = [round(x, 5) for x in sim_xm]
        # --------------------------------------

        sc_mm = [1-df['s1_mm'].mean(),
                 1-df['s2_mm'].mean(),
                 1-df['s3_mm'].mean(),
                 1-df['s4_mm'].mean(),
                 1-df['s5_mm'].mean(),
                 1-df['s6_mm'].mean()
                 ]

        sc_xx = [1-df['s1_xx'].mean(),
                 1-df['s2_xx'].mean(),
                 1-df['s3_xx'].mean(),
                 1-df['s4_xx'].mean(),
                 1-df['s5_xx'].mean(),
                 1-df['s6_xx'].mean()
                 ]

        sc_mx = [1-df['s1_mx'].mean(),
                 1-df['s2_mx'].mean(),
                 1-df['s3_mx'].mean(),
                 1-df['s4_mx'].mean(),
                 1-df['s5_mx'].mean(),
                 1-df['s6_mx'].mean()
                 ]

        sc_xm = [1-df['s1_xm'].mean(),
                 1-df['s2_xm'].mean(),
                 1-df['s3_xm'].mean(),
                 1-df['s4_xm'].mean(),
                 1-df['s5_xm'].mean(),
                 1-df['s6_xm'].mean()
                 ]

        sc_mm_r = [round(x, 5) for x in sc_mm]
        sc_xx_r = [round(x, 5) for x in sc_xx]
        sc_mx_r = [round(x, 5) for x in sc_mx]
        sc_xm_r = [round(x, 5) for x in sc_xm]
        # ------------------------------------

        c_mm_r_sum = sum(sc_mm_r) - 0.7*len(sc_mm_r)
        sum_mm = round(c_mm_r_sum,5)
        c_mm_r_s = [round((x - 0.7)/sum_mm, 5) for x in sc_mm_r]

        print(c_mm_r_s)

        c_xx_r_sum = sum(sc_xx_r) - 0.7*len(sc_xx_r)
        sum_xx = round(c_xx_r_sum,5)
        c_xx_r_s = [round((x - 0.7)/sum_xx, 5) for x in sc_xx_r]

        c_mx_r_sum = sum(sc_mx_r) - 0.7 * len(sc_mx_r)
        sum_mx = round(c_mx_r_sum, 5)
        c_mx_r_s = [round((x - 0.7) / sum_mx, 5) for x in sc_mx_r]

        c_xm_r_sum = sum(sc_xm_r) - 0.7 * len(sc_xm_r)
        sum_xm = round(c_xm_r_sum, 5)
        c_xm_r_s = [round((x - 0.7) / sum_xm, 5) for x in sc_xm_r]

        # ---------------------------------------
        wspd1 = w.create_sheet("Significance", 2)
        wspd1['A1'] = 'sim1_mm'
        wspd1['B1'] = 'sim2_mm'
        wspd1['C1'] = 'sim3_mm'
        wspd1['D1'] = 'sim4_mm'
        wspd1['E1'] = 'sim5_mm'
        wspd1['F1'] = 'sim6_mm'
        wspd1['G1'] = 'SCom_mm'
        wspd1.append(sim_mm_r)

        wspd1.append(['sim1_xx', 'sim2_xx', 'sim3_xx', 'sim4_xx', 'sim5_xx', 'sim6_xx', 'SCom_xx'])
        wspd1.append(sim_xx_r)

        wspd1.append(['sim1_mx', 'sim2_xx', 'sim3_xx', 'sim4_xx', 'sim5_xx', 'sim6_xx''SCom_mx'])
        wspd1.append(sim_mx_r)

        wspd1.append(['sim1_xm', 'sim2_xm', 'sim3_xm', 'sim4_xm', 'sim5_xm', 'sim6_xm''SCom_xm'])
        wspd1.append(sim_xm_r)
        wspd1.append(['', '', '', '', '', ''])

        # --
        wspd1.append(['s1_mm', 's2_mm', 's3_mm', 's4_mm', 's5_mm', 's6_mm'])
        wspd1.append(sc_mm_r)

        wspd1.append(['s1_xx', 's2_xx', 's3_xx', 's4_xx', 's5_xx', 's6_xx'])
        wspd1.append(sc_xx_r)

        wspd1.append(['s1_mx', 's2_xx', 's3_xx', 's4_xx', 's5_xx', 's6_xx'])
        wspd1.append(sc_mx_r)

        wspd1.append(['s1_xm', 's2_xm', 's3_xm', 's4_xm', 's5_xm', 's6_xm'])
        wspd1.append(sc_xm_r)

        wspd1.append(['', '', '', '', '', ''])
        wspd1.append(['sum_mm', '', '', '', '', ''])
        wspd1.append([str(sum_mm), '', '', '', '', ''])

        wspd1.append(['a_mm_factors', '', '', '', '', ''])
        wspd1.append([str(c_mm_r_s), '', '', '', '', ''])

        print(sum_mm)
        print(c_mm_r_s)
        print('-----------')
        print(c_mm_r_s[0])
        print(c_mm_r_s[5])
        #
        wspd1.append(['', '', '', '', '', ''])
        wspd1.append(['sum_xx', '', '', '', '', ''])
        wspd1.append([str(sum_xx), '', '', '', '', ''])
        wspd1.append(['a_xx_factors', '', '', '', '', ''])
        wspd1.append([str(c_xx_r_s), '', '', '', '', ''])
        #
        wspd1.append(['', '', '', '', '', ''])
        wspd1.append(['sum_mx', '', '', '', '', ''])
        wspd1.append([str(sum_mx), '', '', '', '', ''])
        wspd1.append(['a_mx_factors', '', '', '', '', ''])
        wspd1.append([str(c_mx_r_s), '', '', '', '', ''])
        #
        wspd1.append(['', '', '', '', '', ''])
        wspd1.append(['sum_xm', '', '', '', '', ''])
        wspd1.append([str(sum_xm), '', '', '', '', ''])
        wspd1.append(['a_xm_factors', '', '', '', '', ''])
        wspd1.append([str(c_xm_r_s), '', '', '', '', ''])

        dfp = pd.DataFrame()
        for i in range(1, 7):
            smm = 'x'+str(i)+'_min'
            sxx = 'x'+str(i)+'_max'
            if i == 1:
                dfp['min_min'] = c_mm_r_s[i-1]*df[smm]
                dfp['min_max'] = c_mx_r_s[i-1]*df[sxx]
                dfp['max_min'] = c_xm_r_s[i-1]*df[smm]
                dfp['max_max'] = c_xx_r_s[i-1]*df[sxx]
            else:
                dfp['min_min']=dfp['min_min']+c_mm_r_s[i-1]*df[smm]
                dfp['min_max']=dfp['min_max']+c_mx_r_s[i-1]*df[sxx]
                dfp['max_min']=dfp['max_min']+c_xm_r_s[i-1]*df[smm]
                dfp['max_max']=dfp['max_max']+c_xx_r_s[i-1]*df[sxx]

        dfp['max_value'] = dfp.max(axis=1)
        dfp['min_value'] = dfp.min(axis=1)

        wspd2 = w.create_sheet("Potential", 3)
        for r in dataframe_to_rows(dfp, index=False, header=True):
            wspd2.append(r)

        w.save(sr)


def main():
    ll = ['1960-ALL RANGES', '1970-ALL RANGES', '1978-ALL RANGES', '1985-ALL RANGES']
    pd_ = ProcessData(ll)
    pd_.main()
    #pd_.k_years(3, ll)


if __name__ == '__main__':
    main()

